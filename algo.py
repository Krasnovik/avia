from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl import load_workbook, open


#schedule_xlsx = open('data\Schedule_LT.xlsx', read_only=True) 

mct_gruz = 'data\MCT_груз.xlsx'
schedule_xlsx  = 'data\Schedule_LT.xlsx'

wb = load_workbook(schedule_xlsx)
mct = load_workbook(mct_gruz)

answer = {}
city = []

ws = wb.active
m_row = ws.max_row

mct_ws = mct.active
mct_row = mct_ws.max_row

for i in range(2, m_row + 1):
    cell_obj = ws.cell(row = i, column = 2)
    if cell_obj.value not in city:
        city.append(cell_obj.value)

        for a in range(2, m_row + 1):
            cell_obj1 = ws.cell(row = a, column = 2)
            if cell_obj1.value == cell_obj.value:
                g1 = ws.cell(row = a, column = 4)
                answer[cell_obj.value]=[g1.value]
                break
        
        for a in range(2, m_row + 1):
            cell_obj1 = ws.cell(row = a, column = 2)
            if cell_obj1.value == cell_obj.value:
                g1 = ws.cell(row = a, column = 4)
                answer[cell_obj.value].append(g1.value)


def dfs_paths(graph, n, start, goal, path=[], count=0):
    path = path + [start]

    if start == goal and len(path) <= n+2:
        print(path)
        return
    
    for next_node in graph[start]:
        if next_node not in path:
            #проверка на время + время стоянки
            dfs_paths(graph, n, next_node, goal, path, count+1)


graph = {'a':['b'],
         'b':['a', 'c', 'c', 'c'],
         'c':['b', 'b', 'b']
    
}

start = 'KQT'
end = 'OSS'
n = 2

dfs_paths(answer, n , start, end, [], 0)

print("finish")

print("-----------------------------")

from datetime import datetime as dt 

for i in range(2, m_row + 1):
    cell_obj = ws.cell(row = i, column = 3)
    cell_obj1 = ws.cell(row = i, column = 5)
    time_1 = dt.strptime(str(cell_obj.value),"%H:%M:%S")
    time_2 = dt.strptime(str(cell_obj1.value),"%H:%M:%S")
    time_interval = time_2 - time_1
    #print(time_interval)


wb.close()